from urllib import request

#regex procesamiento
import re

from openpyxl import Workbook, load_workbook
import openpyxl.utils.cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

def validacfdi(rfc_emisor, rfc_receptor, total, uuid):
    webservice = 'https://consultaqr.facturaelectronica.sat.gob.mx/consultacfdiservice.svc'
    soap = """<?xml version="1.0" encoding="UTF-8"?>
    <soap:Envelope
        xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/"
        xmlns:xsd="http://www.w3.org/2001/XMLSchema"
        xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
        <soap:Header/>
        <soap:Body>
        <Consulta xmlns="http://tempuri.org/">
            <expresionImpresa>
                ?re="""+rfc_emisor+"""&amp;rr="""+rfc_receptor+"""&amp;tt="""+total+"""&amp;id="""+uuid+"""
            </expresionImpresa>
        </Consulta>
        </soap:Body>
    </soap:Envelope>"""

    data = bytes(soap, 'UTF-8')

    headers = {
        'SOAPAction': '"http://tempuri.org/IConsultaCFDIService/Consulta"',
        'Content-type': 'text/xml; charset="UTF-8"'
    }

    req = request.Request(url=webservice, data=data, method='POST')

    for k, v in headers.items():
        req.add_header(k, v)
    try:
        with request.urlopen(req, timeout=10) as f:
            response = f.read().decode('utf-8')
            result = re.search("(?s)(?<=Estado>).+?(?=</a:)", response).group()
            return result
    except Exception as e:
        print(str(e))
        return 'Error SAT'


if __name__ == '__main__':
    print_hi('Primo Nacho este es un demo primer release :)')

    rfc_emisor = ""
    rfc_receptor = ""
    total = ""
    uuid = ""

    ## abrir un excel y luego leer los datos de las filas
    ## celdas a usar con sus valores se tendrian que pasar por parametros por linea de comando
    colRFCEmisor = 'B'
    colRFCReceptor = 'D'
    colTotal = 'N'
    colUUID = 'O'
    colConsultado = 'P'

    excelFileInput = 'excelInput.xlsx'
    excelFileOutput = 'excelOutput.xlsx'

    wb = load_workbook(excelFileInput)
    ws = wb.active

    clVigente = PatternFill(start_color='3cff00',
                          end_color='3cff00',
                          fill_type='solid')


    clCancelado = PatternFill(start_color='ff1900',
                          end_color='ff1900',
                          fill_type='solid')

    clErrorSat = PatternFill(start_color='fff700',
                          end_color='fff700',
                          fill_type='solid')

    totalFilas = 0
    totalColumnas = 0

    filainicial = 11

    # print the total number of rows
    #print('Total de filas')
    totalFilas = ws.max_row


    #print('Total de columnas')
    totalColumnas = ws.max_column

    for fila in range(filainicial,  totalFilas + 1):
        for columna in range(1,  totalColumnas + 1):

            letradelacolumna = openpyxl.utils.cell.get_column_letter(columna)
            #print('letra de la columna ' + letradelacolumna)

            ##obtener el rfc del emisor
            if colRFCEmisor == letradelacolumna:
                #print('colRFCEmisor celda a con el valor a usar ' + colRFCEmisor + ' ' + letradelacolumna)
                rfc_emisor = ws[colRFCEmisor + str(fila)].value
                #print(rfc_emisor)

            ##obtener el rfc del receptor
            if colRFCReceptor == letradelacolumna:
                #print('colRFCReceptor celda a con el valor a usar ' + colRFCReceptor + ' ' + letradelacolumna)
                rfc_receptor = ws[colRFCReceptor + str(fila)].value
                #print(rfc_receptor)

            ##obtener el total
            if colTotal == letradelacolumna:
                #print('colRFCReceptor celda a con el valor a usar ' + colTotal + ' ' + letradelacolumna)
                total = ws[colTotal + str(fila)].value
                #print(total)

            ##obtener el uuid
            if colUUID == letradelacolumna:
                #print('colRFCReceptor celda a con el valor a usar ' + colUUID + ' ' + letradelacolumna)
                uuid = ws[colUUID + str(fila)].value
                #print(uuid)

            #la letra esla que corresponde con el valor final donde va a estar el estatus
            #se considera que es la ultima letra a la derecha
            if colConsultado == letradelacolumna:
                #print('consuta estatus cfdi')
                #print(fila, columna, ws.cell(row=fila, column=columna).value)
                #print('columna y fila ' + colConsultado + str(fila) )

                ## ya con los datos obtene el estatus en el sat de estos valores

                estadocfdi = validacfdi(rfc_emisor, rfc_receptor, total, uuid)


                if estadocfdi == 'Vigente':
                    #print(estadocfdi)
                    ws[colConsultado + str(fila)] = estadocfdi
                    cellestadocfdi = ws[colConsultado + str(fila)]
                    fuente = Font(color="000000")
                    cellestadocfdi.font = fuente
                    cellestadocfdi.fill = clVigente

                elif estadocfdi == 'Cancelado':
                    #print(estadocfdi)
                    ws[colConsultado + str(fila)] = estadocfdi
                    cellestadocfdi = ws[colConsultado + str(fila)]
                    fuente = Font(color="000000")
                    cellestadocfdi.font = fuente
                    cellestadocfdi.fill = clCancelado

                elif estadocfdi == 'Error SAT':
                    #print(estadocfdi)
                    ws[colConsultado + str(fila)] = estadocfdi
                    cellestadocfdi = ws[colConsultado + str(fila)]
                    fuente = Font(color="000000")
                    cellestadocfdi.font = fuente
                    cellestadocfdi.fill = clErrorSat
                print('Consultando datos de %s %s %s %s el estado del cfdi es : %s' % (
                rfc_emisor, rfc_receptor, total, uuid, estadocfdi))

    wb.save(excelFileOutput)
